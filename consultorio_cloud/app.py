from flask import Flask, render_template, request, jsonify, send_file, abort
import os, base64, io
from datetime import datetime, date
from pathlib import Path

app = Flask(__name__)

# ── BASE DE DATOS: PostgreSQL en nube, SQLite local ────────────────────────────
DATABASE_URL = os.environ.get("DATABASE_URL")
IS_PG = bool(DATABASE_URL)
FOTOS_PATH = Path(os.environ.get("FOTOS_PATH", str(Path(__file__).parent / "instance" / "fotos")))

if IS_PG:
    import psycopg2, psycopg2.extras

def get_conn():
    if IS_PG:
        return psycopg2.connect(DATABASE_URL, sslmode='require')
    else:
        import sqlite3
        db_path = Path(__file__).parent / "instance" / "consultorio.db"
        db_path.parent.mkdir(exist_ok=True)
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

def query(sql, params=(), one=False):
    sql_q = sql.replace("?", "%s") if IS_PG else sql
    conn = get_conn()
    try:
        if IS_PG:
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(sql_q, params)
            if one:
                r = cur.fetchone(); return dict(r) if r else None
            return [dict(r) for r in cur.fetchall()]
        else:
            cur = conn.execute(sql_q, params)
            if one:
                r = cur.fetchone(); return dict(r) if r else None
            return [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()

def execute(sql, params=()):
    sql_q = sql.replace("?", "%s") if IS_PG else sql
    conn = get_conn()
    try:
        if IS_PG:
            cur = conn.cursor(); cur.execute(sql_q, params); conn.commit()
        else:
            conn.execute(sql_q, params); conn.commit()
    finally:
        conn.close()

def insert(sql, params=()):
    sql_q = sql.replace("?", "%s") if IS_PG else sql
    if IS_PG and "RETURNING" not in sql_q.upper():
        sql_q += " RETURNING id"
    conn = get_conn()
    try:
        if IS_PG:
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(sql_q, params); conn.commit()
            r = cur.fetchone(); return r["id"] if r else None
        else:
            cur = conn.execute(sql_q, params); conn.commit(); return cur.lastrowid
    finally:
        conn.close()

def scalar(sql, params=()):
    r = query(sql, params, one=True)
    if r is None: return 0
    return list(r.values())[0] or 0

def init_db():
    FOTOS_PATH.mkdir(parents=True, exist_ok=True)
    if IS_PG:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
        CREATE TABLE IF NOT EXISTS pacientes (
            id SERIAL PRIMARY KEY, nombre TEXT NOT NULL, apellido TEXT NOT NULL,
            dni TEXT UNIQUE, fecha_nacimiento TEXT, telefono TEXT, email TEXT,
            alergias TEXT, notas TEXT, activa INTEGER DEFAULT 1, fecha_alta TEXT DEFAULT CURRENT_DATE);
        CREATE TABLE IF NOT EXISTS tratamientos (
            id SERIAL PRIMARY KEY, nombre TEXT NOT NULL, descripcion TEXT, precio_base REAL DEFAULT 0);
        CREATE TABLE IF NOT EXISTS profesionales (
            id SERIAL PRIMARY KEY, nombre TEXT NOT NULL, especialidad TEXT);
        CREATE TABLE IF NOT EXISTS sesiones (
            id SERIAL PRIMARY KEY, paciente_id INTEGER NOT NULL REFERENCES pacientes(id),
            tratamiento_id INTEGER REFERENCES tratamientos(id),
            profesional_id INTEGER REFERENCES profesionales(id),
            fecha TEXT NOT NULL, hora TEXT, detalle TEXT, producto TEXT,
            costo REAL DEFAULT 0, estado TEXT DEFAULT 'realizada', notas TEXT);
        CREATE TABLE IF NOT EXISTS turnos (
            id SERIAL PRIMARY KEY, paciente_id INTEGER NOT NULL REFERENCES pacientes(id),
            tratamiento_id INTEGER REFERENCES tratamientos(id),
            profesional_id INTEGER REFERENCES profesionales(id),
            fecha TEXT NOT NULL, hora TEXT NOT NULL, duracion INTEGER DEFAULT 60,
            estado TEXT DEFAULT 'pendiente', notas TEXT);
        CREATE TABLE IF NOT EXISTS pagos (
            id SERIAL PRIMARY KEY, paciente_id INTEGER REFERENCES pacientes(id),
            sesion_id INTEGER REFERENCES sesiones(id),
            fecha TEXT NOT NULL DEFAULT CURRENT_DATE, monto REAL NOT NULL,
            concepto TEXT, metodo TEXT DEFAULT 'efectivo', tipo TEXT DEFAULT 'ingreso');
        CREATE TABLE IF NOT EXISTS fotos (
            id SERIAL PRIMARY KEY, paciente_id INTEGER NOT NULL REFERENCES pacientes(id),
            sesion_id INTEGER REFERENCES sesiones(id), etapa TEXT NOT NULL DEFAULT 'antes',
            fecha TEXT DEFAULT CURRENT_DATE, descripcion TEXT, filename TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS presupuestos (
            id SERIAL PRIMARY KEY, paciente_id INTEGER REFERENCES pacientes(id),
            fecha TEXT DEFAULT CURRENT_DATE, validez_dias INTEGER DEFAULT 30,
            observaciones TEXT, estado TEXT DEFAULT 'pendiente', total REAL DEFAULT 0);
        CREATE TABLE IF NOT EXISTS presupuesto_items (
            id SERIAL PRIMARY KEY, presupuesto_id INTEGER NOT NULL REFERENCES presupuestos(id) ON DELETE CASCADE,
            descripcion TEXT NOT NULL, cantidad INTEGER DEFAULT 1,
            precio_unitario REAL DEFAULT 0, subtotal REAL DEFAULT 0);
        """)
        cur.execute("SELECT COUNT(*) FROM tratamientos"); count = cur.fetchone()[0]
        if count == 0:
            cur.execute("""INSERT INTO tratamientos (nombre, precio_base) VALUES
                ('Botox',45000),('Relleno de labios',62000),('Limpieza facial',28000),
                ('Peeling químico',32000),('Depilación láser',38000),('Microblading',55000),
                ('Consulta inicial',15000),('Hidratación profunda',22000),
                ('Mesoterapia facial',35000),('Radiofrecuencia',42000)""")
            cur.execute("""INSERT INTO profesionales (nombre, especialidad) VALUES
                ('Dra. Gómez','Medicina estética'),('Dra. López','Dermatología')""")
        conn.commit(); conn.close()
    else:
        import sqlite3
        db_path = Path(__file__).parent / "instance" / "consultorio.db"
        db_path.parent.mkdir(exist_ok=True)
        conn = sqlite3.connect(db_path)
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS pacientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL, apellido TEXT NOT NULL,
            dni TEXT UNIQUE, fecha_nacimiento TEXT, telefono TEXT, email TEXT,
            alergias TEXT, notas TEXT, activa INTEGER DEFAULT 1, fecha_alta TEXT DEFAULT CURRENT_DATE);
        CREATE TABLE IF NOT EXISTS tratamientos (
            id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL, descripcion TEXT, precio_base REAL DEFAULT 0);
        CREATE TABLE IF NOT EXISTS profesionales (
            id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL, especialidad TEXT);
        CREATE TABLE IF NOT EXISTS sesiones (
            id INTEGER PRIMARY KEY AUTOINCREMENT, paciente_id INTEGER NOT NULL REFERENCES pacientes(id),
            tratamiento_id INTEGER REFERENCES tratamientos(id), profesional_id INTEGER REFERENCES profesionales(id),
            fecha TEXT NOT NULL, hora TEXT, detalle TEXT, producto TEXT,
            costo REAL DEFAULT 0, estado TEXT DEFAULT 'realizada', notas TEXT);
        CREATE TABLE IF NOT EXISTS turnos (
            id INTEGER PRIMARY KEY AUTOINCREMENT, paciente_id INTEGER NOT NULL REFERENCES pacientes(id),
            tratamiento_id INTEGER REFERENCES tratamientos(id), profesional_id INTEGER REFERENCES profesionales(id),
            fecha TEXT NOT NULL, hora TEXT NOT NULL, duracion INTEGER DEFAULT 60,
            estado TEXT DEFAULT 'pendiente', notas TEXT);
        CREATE TABLE IF NOT EXISTS pagos (
            id INTEGER PRIMARY KEY AUTOINCREMENT, paciente_id INTEGER REFERENCES pacientes(id),
            sesion_id INTEGER REFERENCES sesiones(id), fecha TEXT NOT NULL DEFAULT CURRENT_DATE,
            monto REAL NOT NULL, concepto TEXT, metodo TEXT DEFAULT 'efectivo', tipo TEXT DEFAULT 'ingreso');
        CREATE TABLE IF NOT EXISTS fotos (
            id INTEGER PRIMARY KEY AUTOINCREMENT, paciente_id INTEGER NOT NULL REFERENCES pacientes(id),
            sesion_id INTEGER REFERENCES sesiones(id), etapa TEXT NOT NULL DEFAULT 'antes',
            fecha TEXT DEFAULT CURRENT_DATE, descripcion TEXT, filename TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS presupuestos (
            id INTEGER PRIMARY KEY AUTOINCREMENT, paciente_id INTEGER REFERENCES pacientes(id),
            fecha TEXT DEFAULT CURRENT_DATE, validez_dias INTEGER DEFAULT 30,
            observaciones TEXT, estado TEXT DEFAULT 'pendiente', total REAL DEFAULT 0);
        CREATE TABLE IF NOT EXISTS presupuesto_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT, presupuesto_id INTEGER NOT NULL REFERENCES presupuestos(id) ON DELETE CASCADE,
            descripcion TEXT NOT NULL, cantidad INTEGER DEFAULT 1,
            precio_unitario REAL DEFAULT 0, subtotal REAL DEFAULT 0);
        INSERT OR IGNORE INTO tratamientos (nombre, precio_base) VALUES
            ('Botox',45000),('Relleno de labios',62000),('Limpieza facial',28000),
            ('Peeling químico',32000),('Depilación láser',38000),('Microblading',55000),
            ('Consulta inicial',15000),('Hidratación profunda',22000),
            ('Mesoterapia facial',35000),('Radiofrecuencia',42000);
        INSERT OR IGNORE INTO profesionales (nombre, especialidad) VALUES
            ('Dra. Gómez','Medicina estética'),('Dra. López','Dermatología');
        """)
        conn.commit(); conn.close()
    print(f"DB lista: {'PostgreSQL' if IS_PG else 'SQLite local'}")

# ── RUTAS ──────────────────────────────────────────────────────────────────────
@app.route("/")
def index(): return render_template("index.html")

@app.route("/api/pacientes", methods=["GET"])
def get_pacientes():
    q = request.args.get("q","").strip()
    like = "ILIKE" if IS_PG else "LIKE"
    base = """SELECT p.*, COUNT(DISTINCT s.id) as total_sesiones, MAX(s.fecha) as ultima_sesion
              FROM pacientes p LEFT JOIN sesiones s ON s.paciente_id=p.id"""
    if q:
        rows = query(f"{base} WHERE p.nombre {like} ? OR p.apellido {like} ? OR p.dni {like} ? GROUP BY p.id ORDER BY p.apellido",
            (f"%{q}%",f"%{q}%",f"%{q}%"))
    else:
        rows = query(base+" GROUP BY p.id ORDER BY p.apellido")
    return jsonify(rows)

@app.route("/api/pacientes/<int:pid>", methods=["GET"])
def get_paciente(pid):
    p = query("SELECT * FROM pacientes WHERE id=?", (pid,), one=True)
    if not p: return jsonify({"error":"no encontrado"}), 404
    sesiones = query("""SELECT s.*, t.nombre as tratamiento, pr.nombre as profesional
        FROM sesiones s LEFT JOIN tratamientos t ON t.id=s.tratamiento_id
        LEFT JOIN profesionales pr ON pr.id=s.profesional_id
        WHERE s.paciente_id=? ORDER BY s.fecha DESC""", (pid,))
    pagos = query("SELECT * FROM pagos WHERE paciente_id=? ORDER BY fecha DESC", (pid,))
    fotos = query("SELECT * FROM fotos WHERE paciente_id=? ORDER BY etapa,fecha", (pid,))
    pres = query("SELECT * FROM presupuestos WHERE paciente_id=? ORDER BY fecha DESC", (pid,))
    for pr in pres: pr["items"] = query("SELECT * FROM presupuesto_items WHERE presupuesto_id=?",(pr["id"],))
    return jsonify({"paciente":p,"sesiones":sesiones,"pagos":pagos,"fotos":fotos,"presupuestos":pres})

@app.route("/api/pacientes", methods=["POST"])
def crear_paciente():
    d = request.json
    new_id = insert("INSERT INTO pacientes (nombre,apellido,dni,fecha_nacimiento,telefono,email,alergias,notas) VALUES (?,?,?,?,?,?,?,?)",
        (d.get("nombre"),d.get("apellido"),d.get("dni"),d.get("fecha_nacimiento"),
         d.get("telefono"),d.get("email"),d.get("alergias"),d.get("notas")))
    return jsonify({"id":new_id,"ok":True})

@app.route("/api/pacientes/<int:pid>", methods=["PUT"])
def actualizar_paciente(pid):
    d = request.json
    execute("UPDATE pacientes SET nombre=?,apellido=?,dni=?,fecha_nacimiento=?,telefono=?,email=?,alergias=?,notas=?,activa=? WHERE id=?",
        (d.get("nombre"),d.get("apellido"),d.get("dni"),d.get("fecha_nacimiento"),
         d.get("telefono"),d.get("email"),d.get("alergias"),d.get("notas"),d.get("activa",1),pid))
    return jsonify({"ok":True})

@app.route("/api/sesiones", methods=["GET","POST"])
def sesiones():
    if request.method=="POST":
        d = request.json
        new_id = insert("INSERT INTO sesiones (paciente_id,tratamiento_id,profesional_id,fecha,hora,detalle,producto,costo,estado,notas) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (d.get("paciente_id"),d.get("tratamiento_id"),d.get("profesional_id"),
             d.get("fecha",str(date.today())),d.get("hora"),d.get("detalle"),
             d.get("producto"),d.get("costo",0),d.get("estado","realizada"),d.get("notas")))
        return jsonify({"id":new_id,"ok":True})
    pid = request.args.get("paciente_id")
    base = """SELECT s.*, t.nombre as tratamiento, pr.nombre as profesional, p.nombre||' '||p.apellido as paciente_nombre
              FROM sesiones s LEFT JOIN tratamientos t ON t.id=s.tratamiento_id
              LEFT JOIN profesionales pr ON pr.id=s.profesional_id LEFT JOIN pacientes p ON p.id=s.paciente_id"""
    rows = query(base+" WHERE s.paciente_id=? ORDER BY s.fecha DESC",(pid,)) if pid else query(base+" ORDER BY s.fecha DESC LIMIT 50")
    return jsonify(rows)

@app.route("/api/turnos", methods=["GET","POST"])
def turnos():
    if request.method=="POST":
        d = request.json
        new_id = insert("INSERT INTO turnos (paciente_id,tratamiento_id,profesional_id,fecha,hora,duracion,estado,notas) VALUES (?,?,?,?,?,?,?,?)",
            (d.get("paciente_id"),d.get("tratamiento_id"),d.get("profesional_id"),
             d.get("fecha"),d.get("hora"),d.get("duracion",60),d.get("estado","pendiente"),d.get("notas")))
        return jsonify({"id":new_id,"ok":True})
    fecha=request.args.get("fecha"); mes=request.args.get("mes")
    if fecha:
        rows = query("""SELECT tu.*, t.nombre as tratamiento, p.nombre||' '||p.apellido as paciente_nombre, pr.nombre as profesional
            FROM turnos tu LEFT JOIN tratamientos t ON t.id=tu.tratamiento_id
            LEFT JOIN pacientes p ON p.id=tu.paciente_id LEFT JOIN profesionales pr ON pr.id=tu.profesional_id
            WHERE tu.fecha=? ORDER BY tu.hora""",(fecha,))
    elif mes:
        rows = query("SELECT fecha, COUNT(*) as cantidad FROM turnos WHERE fecha LIKE ? GROUP BY fecha",(f"{mes}%",))
    else:
        rows = query("""SELECT tu.*, t.nombre as tratamiento, p.nombre||' '||p.apellido as paciente_nombre
            FROM turnos tu LEFT JOIN tratamientos t ON t.id=tu.tratamiento_id LEFT JOIN pacientes p ON p.id=tu.paciente_id
            WHERE tu.fecha>=CURRENT_DATE ORDER BY tu.fecha,tu.hora LIMIT 20""")
    return jsonify(rows)

@app.route("/api/turnos/<int:tid>", methods=["PUT"])
def actualizar_turno(tid):
    execute("UPDATE turnos SET estado=? WHERE id=?",(request.json.get("estado"),tid))
    return jsonify({"ok":True})

@app.route("/api/pagos", methods=["GET","POST"])
def pagos():
    if request.method=="POST":
        d = request.json
        new_id = insert("INSERT INTO pagos (paciente_id,sesion_id,fecha,monto,concepto,metodo,tipo) VALUES (?,?,?,?,?,?,?)",
            (d.get("paciente_id"),d.get("sesion_id"),d.get("fecha",str(date.today())),
             d.get("monto"),d.get("concepto"),d.get("metodo","efectivo"),d.get("tipo","ingreso")))
        return jsonify({"id":new_id,"ok":True})
    mes = request.args.get("mes",str(date.today())[:7])
    movs = query("""SELECT pg.*, p.nombre||' '||p.apellido as paciente_nombre
        FROM pagos pg LEFT JOIN pacientes p ON p.id=pg.paciente_id
        WHERE pg.fecha LIKE ? ORDER BY pg.fecha DESC LIMIT 50""",(f"{mes}%",))
    res = query("""SELECT COALESCE(SUM(CASE WHEN tipo='ingreso' THEN monto ELSE 0 END),0) as ingresos,
        COALESCE(SUM(CASE WHEN tipo='egreso' THEN monto ELSE 0 END),0) as egresos,
        COUNT(CASE WHEN tipo='ingreso' THEN 1 END) as count_ingresos
        FROM pagos WHERE fecha LIKE ?""",(f"{mes}%",),one=True)
    return jsonify({"movimientos":movs,"resumen":res})

@app.route("/api/fotos", methods=["POST"])
def subir_foto():
    d = request.json; pid=d.get("paciente_id"); img=d.get("imagen","")
    if not pid or not img: return jsonify({"error":"faltan datos"}),400
    ext = "png" if "png" in img[:30] else "jpg"
    fname = f"pac{pid}_{d.get('etapa','antes')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
    if "," in img: img = img.split(",")[1]
    (FOTOS_PATH/fname).write_bytes(base64.b64decode(img))
    new_id = insert("INSERT INTO fotos (paciente_id,sesion_id,etapa,fecha,descripcion,filename) VALUES (?,?,?,?,?,?)",
        (pid,d.get("sesion_id"),d.get("etapa","antes"),str(date.today()),d.get("descripcion",""),fname))
    return jsonify({"id":new_id,"ok":True,"filename":fname})

@app.route("/api/fotos/img/<filename>")
def get_foto(filename):
    fp = FOTOS_PATH/filename
    if not fp.exists(): abort(404)
    ext = filename.rsplit(".",1)[-1].lower()
    return send_file(fp, mimetype={"jpg":"image/jpeg","jpeg":"image/jpeg","png":"image/png"}.get(ext,"image/jpeg"))

@app.route("/api/fotos/<int:fid>", methods=["DELETE"])
def del_foto(fid):
    f = query("SELECT * FROM fotos WHERE id=?",(fid,),one=True)
    if f:
        fp = FOTOS_PATH/f["filename"]
        if fp.exists(): fp.unlink()
        execute("DELETE FROM fotos WHERE id=?",(fid,))
    return jsonify({"ok":True})

@app.route("/api/presupuestos", methods=["GET","POST"])
def presupuestos():
    if request.method=="POST":
        d=request.json; items=d.get("items",[])
        total=sum(i.get("cantidad",1)*i.get("precio_unitario",0) for i in items)
        pr_id=insert("INSERT INTO presupuestos (paciente_id,fecha,validez_dias,observaciones,estado,total) VALUES (?,?,?,?,?,?)",
            (d.get("paciente_id"),d.get("fecha",str(date.today())),d.get("validez_dias",30),
             d.get("observaciones",""),d.get("estado","pendiente"),total))
        for it in items:
            insert("INSERT INTO presupuesto_items (presupuesto_id,descripcion,cantidad,precio_unitario,subtotal) VALUES (?,?,?,?,?)",
                (pr_id,it.get("descripcion"),it.get("cantidad",1),it.get("precio_unitario",0),
                 it.get("cantidad",1)*it.get("precio_unitario",0)))
        return jsonify({"id":pr_id,"ok":True})
    pid=request.args.get("paciente_id")
    rows = query("SELECT pr.* FROM presupuestos pr WHERE pr.paciente_id=? ORDER BY pr.fecha DESC",(pid,)) if pid else \
        query("""SELECT pr.*, p.nombre||' '||p.apellido as paciente_nombre FROM presupuestos pr
            LEFT JOIN pacientes p ON p.id=pr.paciente_id ORDER BY pr.fecha DESC LIMIT 30""")
    for r in rows: r["items"]=query("SELECT * FROM presupuesto_items WHERE presupuesto_id=?",(r["id"],))
    return jsonify(rows)

@app.route("/api/presupuestos/<int:prid>", methods=["GET","PUT"])
def presupuesto_det(prid):
    if request.method=="PUT":
        execute("UPDATE presupuestos SET estado=? WHERE id=?",(request.json.get("estado"),prid))
        return jsonify({"ok":True})
    pr=query("""SELECT pr.*, p.nombre||' '||p.apellido as paciente_nombre, p.telefono as paciente_telefono, p.dni as paciente_dni
        FROM presupuestos pr LEFT JOIN pacientes p ON p.id=pr.paciente_id WHERE pr.id=?""",(prid,),one=True)
    if not pr: return jsonify({"error":"no encontrado"}),404
    return jsonify({"presupuesto":pr,"items":query("SELECT * FROM presupuesto_items WHERE presupuesto_id=?",(prid,))})

@app.route("/api/dashboard")
def dashboard():
    hoy=str(date.today()); mes=hoy[:7]
    stats={
        "pacientes_activas": scalar("SELECT COUNT(*) FROM pacientes WHERE activa=1"),
        "turnos_mes": scalar("SELECT COUNT(*) FROM turnos WHERE fecha LIKE ?",(f"{mes}%",)),
        "ingresos_mes": scalar("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE fecha LIKE ? AND tipo='ingreso'",(f"{mes}%",)),
        "turnos_hoy": scalar("SELECT COUNT(*) FROM turnos WHERE fecha=?",(hoy,)),
    }
    turnos_hoy=query("""SELECT tu.*, p.nombre||' '||p.apellido as paciente_nombre, t.nombre as tratamiento
        FROM turnos tu LEFT JOIN pacientes p ON p.id=tu.paciente_id
        LEFT JOIN tratamientos t ON t.id=tu.tratamiento_id WHERE tu.fecha=? ORDER BY tu.hora""",(hoy,))
    nl = "NULLS LAST" if not IS_PG else ""
    ultimas=query(f"""SELECT p.id, p.nombre, p.apellido, MAX(s.fecha) as ultima_sesion, t.nombre as ultimo_tratamiento
        FROM pacientes p LEFT JOIN sesiones s ON s.paciente_id=p.id LEFT JOIN tratamientos t ON t.id=s.tratamiento_id
        GROUP BY p.id, p.nombre, p.apellido ORDER BY ultima_sesion DESC {nl} LIMIT 5""")
    return jsonify({"stats":stats,"turnos_hoy":turnos_hoy,"ultimas_pacientes":ultimas})

@app.route("/api/tratamientos", methods=["GET","POST"])
def trats():
    if request.method=="POST":
        d=request.json
        new_id=insert("INSERT INTO tratamientos (nombre,descripcion,precio_base) VALUES (?,?,?)",
            (d["nombre"],d.get("descripcion",""),float(d.get("precio_base",0))))
        return jsonify({"id":new_id,"ok":True})
    return jsonify(query("SELECT * FROM tratamientos ORDER BY nombre"))

@app.route("/api/tratamientos/<int:tid>", methods=["PUT","DELETE"])
def trat_det(tid):
    if request.method=="DELETE":
        u=scalar("SELECT COUNT(*) FROM sesiones WHERE tratamiento_id=?",(tid,))
        u+=scalar("SELECT COUNT(*) FROM turnos WHERE tratamiento_id=?",(tid,))
        if u>0: return jsonify({"error":f"Tiene {u} registro(s). No se puede eliminar."}),409
        execute("DELETE FROM tratamientos WHERE id=?",(tid,)); return jsonify({"ok":True})
    d=request.json
    execute("UPDATE tratamientos SET nombre=?,descripcion=?,precio_base=? WHERE id=?",
        (d.get("nombre"),d.get("descripcion",""),float(d.get("precio_base",0)),tid))
    return jsonify({"ok":True})

@app.route("/api/profesionales", methods=["GET","POST"])
def profs():
    if request.method=="POST":
        d=request.json
        new_id=insert("INSERT INTO profesionales (nombre,especialidad) VALUES (?,?)",(d["nombre"],d.get("especialidad","")))
        return jsonify({"id":new_id,"ok":True})
    return jsonify(query("SELECT * FROM profesionales ORDER BY nombre"))

@app.route("/api/profesionales/<int:pid>", methods=["PUT","DELETE"])
def prof_det(pid):
    if request.method=="DELETE":
        u=scalar("SELECT COUNT(*) FROM sesiones WHERE profesional_id=?",(pid,))
        u+=scalar("SELECT COUNT(*) FROM turnos WHERE profesional_id=?",(pid,))
        if u>0: return jsonify({"error":f"Tiene {u} registro(s). No se puede eliminar."}),409
        execute("DELETE FROM profesionales WHERE id=?",(pid,)); return jsonify({"ok":True})
    d=request.json
    execute("UPDATE profesionales SET nombre=?,especialidad=? WHERE id=?",(d.get("nombre"),d.get("especialidad",""),pid))
    return jsonify({"ok":True})

def make_headers(ws, headers):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        fill=PatternFill("solid",fgColor="8B4513")
        for ci,h in enumerate(headers,1):
            cell=ws.cell(1,ci,h); cell.fill=fill; cell.font=Font(bold=True,color="FFFFFF")
            cell.alignment=Alignment(horizontal="center")
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=20
    except: pass

@app.route("/api/exportar/pacientes")
def exp_pac():
    try: import openpyxl
    except: return jsonify({"error":"pip install openpyxl"}),500
    rows=query("""SELECT p.nombre, p.apellido, p.dni, p.fecha_nacimiento, p.telefono, p.email,
        p.alergias, p.fecha_alta, COUNT(DISTINCT s.id) as sesiones, MAX(s.fecha) as ultima_sesion
        FROM pacientes p LEFT JOIN sesiones s ON s.paciente_id=p.id GROUP BY p.id ORDER BY p.apellido""")
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Pacientes"
    make_headers(ws,["Nombre","Apellido","DNI","Nacimiento","Teléfono","Email","Alergias","Alta","Sesiones","Última visita"])
    for r in rows: ws.append(list(r.values()))
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name=f"pacientes_{date.today()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/exportar/caja")
def exp_caja():
    try: import openpyxl
    except: return jsonify({"error":"pip install openpyxl"}),500
    mes=request.args.get("mes",str(date.today())[:7])
    rows=query("""SELECT pg.fecha, p.nombre||' '||p.apellido, pg.concepto, pg.metodo, pg.tipo, pg.monto
        FROM pagos pg LEFT JOIN pacientes p ON p.id=pg.paciente_id WHERE pg.fecha LIKE ? ORDER BY pg.fecha""",(f"{mes}%",))
    wb=openpyxl.Workbook(); ws=wb.active; ws.title=f"Caja {mes}"
    make_headers(ws,["Fecha","Paciente","Concepto","Método","Tipo","Monto"])
    for r in rows: ws.append(list(r.values()))
    ws.append([]); ws.append(["","","","","Ingresos:",sum(r["monto"] for r in rows if r["tipo"]=="ingreso")])
    ws.append(["","","","","Egresos:",sum(r["monto"] for r in rows if r["tipo"]=="egreso")])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name=f"caja_{mes}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/exportar/historial")
def exp_hist():
    try: import openpyxl
    except: return jsonify({"error":"pip install openpyxl"}),500
    rows=query("""SELECT s.fecha, p.nombre||' '||p.apellido, t.nombre, pr.nombre, s.producto, s.costo, s.estado, s.notas
        FROM sesiones s LEFT JOIN pacientes p ON p.id=s.paciente_id
        LEFT JOIN tratamientos t ON t.id=s.tratamiento_id LEFT JOIN profesionales pr ON pr.id=s.profesional_id
        ORDER BY s.fecha DESC""")
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Historial"
    make_headers(ws,["Fecha","Paciente","Tratamiento","Profesional","Producto","Costo","Estado","Notas"])
    for r in rows: ws.append(list(r.values()))
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name=f"historial_{date.today()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    init_db()
    if not IS_PG:
        import webbrowser, threading
        threading.Timer(1.2, lambda: webbrowser.open("http://localhost:5000")).start()
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port)
